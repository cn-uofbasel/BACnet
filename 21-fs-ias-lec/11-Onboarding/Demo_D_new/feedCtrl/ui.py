#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 19 15:15:14 2020

@author: yannickrummele
"""

from tkinter import *
from tkinter import ttk
# from uiFunctionsHandler import UiFunctionHandler
from .uiFunctionsHandler import *

from openpyxl import load_workbook, Workbook

import bluetooth
import os

ufh = UiFunctionHandler()


def update_feedIds():
    print("update feedIDs")
    feedIDsTree.delete(*feedIDsTree.get_children())
    allMasterIDs = list(ufh.get_master_ids())

    for i in range(len(allMasterIDs)):
        treeUsername = ufh.get_username(allMasterIDs[i])
        treeID = allMasterIDs[i].hex()
        feedIDsTree.insert('', 'end', treeUsername, text=treeUsername, values=treeID)
        feedIDsTree.item(treeUsername, tags=('master'))

        feedIDs = ufh.get_all_master_ids_feed_ids(allMasterIDs[i])

        trusted = list(ufh.get_trusted())
        blocked = list(ufh.get_blocked())
        for feedID in feedIDs:

            if feedID in trusted:
                treeApplicationName = ufh.get_application(feedID)
                treeApplicationID = feedID.hex()
                if treeApplicationName is not None:
                    followedChildname = treeUsername + " followed " + treeApplicationName
                    feedIDsTree.insert(treeUsername, 'end', followedChildname, text=treeApplicationName,
                                       values=treeApplicationID)
                    feedIDsTree.item(followedChildname, tags=('followed'))
            elif feedID in blocked:
                treeApplicationName = ufh.get_application(feedID)
                treeApplicationID = feedID.hex()
                if treeApplicationName is not None:
                    blockedChildname = treeUsername + " blocked " + treeApplicationName
                    feedIDsTree.insert(treeUsername, 'end', blockedChildname, text=treeApplicationName,
                                       values=treeApplicationID)
                    feedIDsTree.item(blockedChildname, tags=('blocked'))
            else:
                treeApplicationName = ufh.get_application(feedID)
                treeApplicationID = feedID.hex()
                if treeApplicationName is not None:
                    blockedChildname = treeUsername + " blocked " + treeApplicationName
                    feedIDsTree.insert(treeUsername, 'end', blockedChildname, text=treeApplicationName,
                                       values=treeApplicationID)
                    feedIDsTree.item(blockedChildname)


def setTrusted():
    feedIDsTree.bind('<<TreeviewSelect>>', callback)
    print("trusted")
    curItemID = (feedIDsTree.selection()[0])
    curItem = "".join(feedIDsTree.item(feedIDsTree.selection()[0])['values'])
    curID = bytes.fromhex(curItem)
    print(curItem)
    print(curID)
    ufh.set_trusted(curID, True)
    feedIDsTree.item(curItemID, tags=('followed'))


def setUntrusted():
    feedIDsTree.bind('<<TreeviewSelect>>', callback)
    print("untrusted")
    curItemID = (feedIDsTree.selection()[0])
    curItem = "".join(feedIDsTree.item(feedIDsTree.selection()[0])['values'])
    curID = bytes.fromhex(curItem)
    print(curItem)
    print(curID)
    ufh.set_trusted(curID, False)
    feedIDsTree.item(curItemID, tags=('blocked'))


def callback(event):
    print(feedIDsTree.selection()[0])


def updateUsername():
    ufh.set_username(entryUsername.get())
    print("New username: " + entryUsername.get())
    # entryUsername.delete(0,END)


def updateRadius():
    try:
        radius = int(entryRadius.get())
        ufh.set_radius(radius)
        print("New radius: " + entryRadius.get())

    except Exception as e:
        print("Insert a integer for the radius")
        print(e)
        print(entryRadius.get() + " is not a integer")


def onboarding():
    child = Toplevel()
    child.title("BACnet Onboarding")
    Label(child, text="Your starting the onboarding process to meet new people").grid(row=0)
    Label(child,
          text="Choose between server or client, your partner needs to press the opposite. You both need to have bluetooth on").grid(
        row=1)
    Button(child, text="Client", width=25, command=clientBT).grid(row=3)
    Button(child, text="Server", width=25, command=serverBT).grid(row=4)
    Button(child, text='Close', width=25, command=child.destroy).grid(row=5)


def serverBT():
    if not os.path.isfile('./data.xlsx'):
        workbook = Workbook()  # creates new workbook
        workbook.save('data.xlsx')  # saves the workbook
    workbook = load_workbook("data.xlsx")
    worksheet = workbook.worksheets[0]

    srvSock = bluetooth.BluetoothSocket(bluetooth.RFCOMM)
    srvSock.bind(("", bluetooth.PORT_ANY))
    srvSock.listen(1)

    port = srvSock.getsockname()[1]

    uuid = "94f39d29-7d6d-437d-973b-fba39e49d4ee"

    bluetooth.advertise_service(srvSock, "SampleServer", service_id=uuid,
                                service_classes=[uuid, bluetooth.SERIAL_PORT_CLASS],
                                profiles=[bluetooth.SERIAL_PORT_PROFILE],
                                )

    print("Waiting for connection on RFCOMM port", port)

    cliSock, cliInfo = srvSock.accept()
    print("Connection from", cliInfo)

    try:
        while True:
            partnerID = cliSock.recv(1024)
            worksheet.cell(column=1, row=worksheet.max_row + 1, value=partnerID)
            workbook.save('data.xlsx')
            cliSock.send(str(ufh.get_host_master_id()))
            if not partnerID:
                break
            print("Received", partnerID)
    except OSError:
        pass

    print("Finished")
    cliSock.close()
    srvSock.close()


def clientBT():
    if not os.path.isfile('./data.xlsx'):
        workbook = Workbook()
        workbook.save('data.xlsx')
    workbook = load_workbook("data.xlsx")
    worksheet = workbook.worksheets[0]


    partnerAddress = None
    while partnerAddress is None:
        partnerAddress = input("enter the name of your bluetooth partner: ")

        nearby_devices = bluetooth.discover_devices()

        for bdaddr in nearby_devices:
            if partnerAddress == bluetooth.lookup_name(bdaddr):
                partnerAddress = bdaddr
                break

        if partnerAddress is not None:
            print("found partner bluetooth device", partnerAddress)
        else:
            print("could not find the device. Try again.")

    print(str(ufh.get_host_master_id()))
    port = 4
    sock = bluetooth.BluetoothSocket(bluetooth.RFCOMM)
    sock.connect((partnerAddress, port))
    sock.send(str(ufh.get_host_master_id()))
    partnerID = sock.recv(1024)
    print("Received", partnerID)
    worksheet.cell(column=1, row=worksheet.max_row + 1, value=partnerID)
    workbook.save('data.xlsx')

    print("Finished")
    sock.close()


# generate_test_data()

root = Tk()
root.title("BACnet feedCtrl")
labelWelcome = Label(root, text="Welcome to BACnet feedCtrl.").grid(row=0)

labelInstruction = Label(root, text="Press the button to update the list of the FeedIDs").grid(row=1)
feedIDsTree = ttk.Treeview(root)
feedIDsTree.grid(row=4)
feedIDsTree.config(columns=('id'))
feedIDsTree.heading('#0', text='Name')
feedIDsTree.heading('id', text='ID')
feedIDsTree.tag_configure('master', background='yellow')
feedIDsTree.tag_configure('blocked', background='red')
feedIDsTree.tag_configure('followed', background='green')
feedIDsTree.config(selectmode='browse')

buttonUpdateFeedIDs = Button(root, text="UpdateFeedIDs", width=25, command=update_feedIds).grid(row=3)

buttonTrust = Button(root, text="Trust", width=25, command=setTrusted).grid(row=5)
buttonUntrust = Button(root, text="Untrust", width=25, command=setUntrusted).grid(row=6)

entryUsername = Entry(root)
entryUsername.grid(row=7)
buttonUpdateUsername = Button(root, text="Update Username", width=25, command=updateUsername).grid(row=8)

entryRadius = Entry(root)
entryRadius.grid(row=9)
buttonUpdateRadius = Button(root, text="Update Radius", width=25, command=updateRadius).grid(row=10)

buttonOnboarding = Button(root, text="Bluetooth-Onboarding", width=25, command=onboarding).grid(row=12)
buttonQuit = Button(root, text='Quit', width=25, command=root.destroy).grid(row=14)


def run():
    try:
        root.mainloop()
        root.destroy()
        root.close()
        exit()
    except:
        pass
